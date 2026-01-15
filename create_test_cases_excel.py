#!/usr/bin/env python3
"""
Script to generate Manual Test Cases Excel file for materials/add.php
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
category_font = Font(bold=True, color="FFFFFF", size=12)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ["TC ID", "Category", "Description", "Pre-conditions", "Test Steps", "Expected Result", "Actual Result", "Status", "Notes"]
ws.append(headers)

# Style header row
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# Test cases data
test_cases = [
    # TC001: Initial Page Load
    ("TC001.1", "Initial Page Load", "โหลดหน้าเพจครั้งแรก",
     "- เข้าสู่ระบบแล้ว\n- Database มีข้อมูลพื้นฐาน",
     "1. เปิดเบราว์เซอร์\n2. เข้า URL: /pages/materials/add.php\n3. รอหน้าเพจโหลดเสร็จ",
     "✅ แสดงหน้าเพจสมบูรณ์\n✅ Breadcrumb: หน้าหลัก > รายการวัสดุ > เพิ่มวัสดุใหม่\n✅ Header แสดง 'เพิ่มข้อมูล Material'\n✅ SSP Preview แสดง '-- เลือกข้อมูลเพื่อดูตัวอย่าง --'\n✅ Progress Steps แสดงขั้นตอน 3 ขั้น\n✅ ไม่มี error ใน Console"),

    ("TC001.2", "Initial Page Load", "โหลด Dropdown ประเภทวัสดุ",
     "หน้าเพจโหลดเสร็จแล้ว",
     "1. คลิก dropdown 'ประเภทวัสดุ'\n2. ดูรายการที่แสดง",
     "✅ แสดงตัวเลือก 'เลือกประเภท' เป็นตัวเลือกแรก\n✅ แสดงรายการประเภทวัสดุทั้งหมดจาก DB\n✅ รูปแบบ: 'ชื่อประเภท (รหัส)' เช่น 'Raw Material (R)'\n✅ มีข้อมูลอย่างน้อย 1 รายการ"),

    ("TC001.3", "Initial Page Load", "โหลด Dropdown กลุ่มวัสดุ",
     "หน้าเพจโหลดเสร็จแล้ว",
     "1. คลิก dropdown 'กลุ่มวัสดุ'\n2. ดูรายการที่แสดง",
     "✅ แสดงตัวเลือก 'เลือกกลุ่ม' เป็นตัวเลือกแรก\n✅ แสดงรายการกลุ่มวัสดุทั้งหมดจาก DB\n✅ รูปแบบ: 'ชื่อกลุ่ม (รหัส)' เช่น 'กระดาษคาร์ตัน (501)'\n✅ มีข้อมูลอย่างน้อย 1 รายการ"),

    ("TC001.4", "Initial Page Load", "โหลด Dropdown ผู้ขาย (Select2)",
     "หน้าเพจโหลดเสร็จแล้ว",
     "1. คลิก dropdown 'ผู้ขาย'\n2. ดูรายการที่แสดง\n3. สังเกตรูปแบบ Select2",
     "✅ แสดง placeholder 'เลือกหรือค้นหาผู้ขาย...'\n✅ แสดงไอคอนลูกศรและช่องค้นหา (Select2 style)\n✅ แสดงรายการผู้ขายทั้งหมดจาก DB\n✅ รูปแบบ: 'ชื่อผู้ขาย (รหัส)'\n✅ มีข้อมูลอย่างน้อย 1 รายการ"),

    ("TC001.5", "Initial Page Load", "โหลด Dropdown หน่วย",
     "หน้าเพจโหลดเสร็จแล้ว",
     "1. คลิก dropdown 'หน่วยหลัก'\n2. ดูรายการที่แสดง",
     "✅ แสดงตัวเลือก 'เลือกหน่วย' เป็นตัวเลือกแรก\n✅ แสดงรายการหน่วยทั้งหมดจาก DB\n✅ รูปแบบ: 'ชื่อหน่วย (สัญลักษณ์)'\n✅ มีข้อมูลอย่างน้อย 1 รายการ"),

    # TC002: SSP Code Preview
    ("TC002.1", "SSP Code Preview", "Preview เมื่อไม่เลือกข้อมูล",
     "โหลดหน้าเพจแล้ว",
     "1. ดู SSP Preview ที่ header",
     "✅ แสดงข้อความ '-- เลือกข้อมูลเพื่อดูตัวอย่าง --'\n✅ ไม่แสดงรหัส SSP"),

    ("TC002.2", "SSP Code Preview", "Preview เมื่อเลือกไม่ครบ",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'ประเภทวัสดุ' = 'Raw Material (R)'\n2. ดู SSP Preview",
     "✅ แสดง '-----' หรือข้อความเดิม\n✅ ยังไม่สร้าง SSP Code"),

    ("TC002.3", "SSP Code Preview", "Preview เมื่อเลือกครบ 3 ฟิลด์",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'ประเภทวัสดุ' = 'Raw Material (R)'\n2. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n3. เลือก 'ผู้ขาย' = 'ABC Company (001)'\n4. ดู SSP Preview\n5. เปิด Console ดู Network tab",
     "✅ เรียก API generate_ssp_code.php ผ่าน POST\n✅ แสดง SSP Code รูปแบบ: R501001XXXXX (12 หลัก)\n✅ Code มีโครงสร้าง: Type(1) + Group(3) + Supplier(3) + RunNumber(5)\n✅ แสดงใน Preview box พร้อมสีพื้นหลัง"),

    ("TC002.4", "SSP Code Preview", "Preview เปลี่ยนแปลงเมื่อเปลี่ยนผู้ขาย",
     "เลือกข้อมูลครบ 3 ฟิลด์แล้ว, มี SSP Code แสดงอยู่",
     "1. จด SSP Code เดิม\n2. เปลี่ยน 'ผู้ขาย' เป็นรายอื่น\n3. ดู SSP Preview",
     "✅ SSP Code เปลี่ยนไป (ตัวเลข 3 ตัวตรงกลางเปลี่ยน)\n✅ อัปเดตแบบ real-time\n✅ ไม่ต้อง refresh หน้า"),

    ("TC002.5", "SSP Code Preview", "Preview เมื่อ API ล้มเหลว",
     "ปิด API server หรือแก้ไข URL API ให้ผิด",
     "1. เลือกข้อมูลครบ 3 ฟิลด์\n2. ดู SSP Preview\n3. ดู Console",
     "✅ แสดง '-----'\n✅ ไม่มี error popup\n✅ Console แสดง AJAX error"),

    # TC003: Form Visibility - กลุ่มกระดาษ
    ("TC003.1", "Form Visibility", "แสดงฟอร์มกระดาษคาร์ตัน (501)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n2. scroll ดูฟอร์มทั้งหมด",
     "✅ แสดง Section 'รายละเอียดเฉพาะ - Paperboard/กระดาษ'\n✅ มีฟิลด์: ประเภทกระดาษ (ไทย/อังกฤษ)\n✅ มีฟิลด์: แบรนด์ 1, แบรนด์ 2, GSM, Caliper\n✅ มีฟิลด์: ขนาด (มม./นิ้ว)\n✅ มีฟิลด์: Laminated 1, 2, การรับรอง\n✅ มีการคำนวณน้ำหนักต่อแผ่น\n✅ ชื่อวัสดุเป็น readonly (พื้นสีเทา)\n✅ แสดง hint 'สำหรับกระดาษ: ระบบจะสร้างชื่ออัตโนมัติ...'"),

    ("TC003.2", "Form Visibility", "แสดงฟอร์มกระดาษอาร์ต (551)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กระดาษอาร์ต (551)'",
     "✅ แสดง Paperboard Section\n✅ พฤติกรรมเหมือน TC003.1"),

    ("TC003.3", "Form Visibility", "แสดงฟอร์มกระดาษอื่นๆ (801-804)",
     "โหลดหน้าเพจแล้ว",
     "1. ทดสอบทีละกลุ่ม: 801, 802, 803, 804\n2. ดูฟอร์มที่แสดง",
     "✅ ทุกกลุ่มแสดง Paperboard Section\n✅ ชื่อวัสดุเป็น readonly"),

    # TC004: Form Visibility - กลุ่มอื่นๆ
    ("TC004.1", "Form Visibility", "แสดงฟอร์มหมึก (553)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'หมึก (553)'",
     "✅ แสดง Section 'รายละเอียดเฉพาะ - Ink/หมึก'\n✅ ซ่อน Paperboard Section\n✅ มีฟิลด์: ประเภทหมึก, สี, กลุ่มหมึก, ด้าน\n✅ มีฟิลด์: แบรนด์กระดาษ, ประเภทกระดาษ, GSM กระดาษ\n✅ มีฟิลด์: Laminated 1, 2, Coating 1, 2\n✅ ชื่อวัสดุไม่เป็น readonly (กรอกได้)"),

    ("TC004.2", "Form Visibility", "แสดงฟอร์มกาว (556)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กาว (556)'",
     "✅ แสดง Section 'รายละเอียดเฉพาะ - Adhesive/กาว'\n✅ มีฟิลด์: ประเภทกาว, ใช้กับวัสดุ, การใช้งาน\n✅ ชื่อวัสดุกรอกได้"),

    ("TC004.3", "Form Visibility", "แสดงฟอร์มเคลือบ (557)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'เคลือบ (557)'",
     "✅ แสดง Section 'รายละเอียดเฉพาะ - Coating/เคลือบผิว'\n✅ มีฟิลด์: ฐานเคลือบ, ประเภท, เอฟเฟกต์, ความหนา\n✅ ชื่อวัสดุกรอกได้"),

    ("TC004.4", "Form Visibility", "แสดงฟอร์มฟิล์ม (714)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'ฟิล์ม (714)'",
     "✅ แสดง Section 'รายละเอียดเฉพาะ - Film/ฟิล์ม'\n✅ มีฟิลด์: ประเภทฟิล์ม, รหัสฟิล์ม, เอฟเฟกต์, ความหนา\n✅ ชื่อวัสดุกรอกได้"),

    ("TC004.5", "Form Visibility", "แสดงฟอร์มฟอยล์ (555)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'ฟอยล์ (555)'",
     "✅ แสดง Section 'รายละเอียดเฉพาะ - Foil/ฟอยล์'\n✅ มีฟิลด์: รหัสฟอยล์, สี, เอฟเฟกต์\n✅ มีฟิลด์ขนาด: กว้าง (มม.), ยาว (เมตร), พื้นที่ (ตร.ม.)\n✅ พื้นที่คำนวณอัตโนมัติ (readonly)\n✅ ชื่อวัสดุกรอกได้"),

    ("TC004.6", "Form Visibility", "แสดงฟอร์มกล่องลูกฟูก (260)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กล่องลูกฟูก (260)'",
     "✅ แสดง Section 'รายละเอียดเฉพาะ - Corrugated Box/กล่องลูกฟูก'\n✅ มีฟิลด์: Case Number\n✅ มีฟิลด์ขนาดภายนอก: กว้าง, ยาว, สูง (มม.)\n✅ มีฟิลด์ขนาดภายใน: กว้าง, ยาว, สูง (มม.)\n✅ มีฟิลด์โครงสร้าง: ประเภทฟลูต, จำนวนชั้น, น้ำหนัก, โลโก้\n✅ มีฟิลด์: Liner 1-3, Flute 1-2\n✅ ชื่อวัสดุกรอกได้"),

    ("TC004.7", "Form Visibility", "สลับกลุ่มไปมา",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n2. ดูฟอร์มที่แสดง\n3. เปลี่ยนเป็น 'หมึก (553)'\n4. ดูฟอร์มที่แสดง\n5. เปลี่ยนกลับเป็น 'กระดาษคาร์ตัน (501)'",
     "✅ ขั้นที่ 2: แสดง Paperboard Section, ซ่อนฟอร์มอื่น\n✅ ขั้นที่ 4: ซ่อน Paperboard, แสดง Ink Section\n✅ ขั้นที่ 5: แสดง Paperboard, ซ่อน Ink Section\n✅ ไม่มี delay หรือ flicker"),

    # TC005: Unit Filtering
    ("TC005.1", "Unit Filtering", "หน่วยเริ่มต้น (ไม่เลือกกลุ่ม)",
     "โหลดหน้าเพจแล้ว",
     "1. คลิก dropdown 'หน่วยหลัก'\n2. นับจำนวนตัวเลือก",
     "✅ แสดงหน่วยทั้งหมดจาก database\n✅ ไม่มีการกรอง"),

    ("TC005.2", "Unit Filtering", "กรองหน่วยสำหรับกระดาษ (501)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n2. คลิก dropdown 'หน่วยหลัก'\n3. ดูรายการที่แสดง",
     "✅ แสดงเฉพาะหน่วย: แผ่น/sheet, ห่อ/pack, รีม/ream, ม้วน/roll, กิโล/กิโลกรัม/kg/kilogram, ตัน/ton/metric ton/t\n✅ ไม่แสดงหน่วยอื่น (เช่น กระป๋อง, ลิตร, กล่อง)\n✅ มีตัวเลือก 'แผ่น' ถูกเลือกอัตโนมัติ (default)"),

    ("TC005.3", "Unit Filtering", "กรองหน่วยสำหรับหมึก (553)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'หมึก (553)'\n2. คลิก dropdown 'หน่วยหลัก'\n3. ดูรายการที่แสดง",
     "✅ แสดงเฉพาะหน่วย: กระป๋อง/can, กิโล/กิโลกรัม/kg/kilogram\n✅ ไม่แสดงหน่วยอื่น (เช่น แผ่น, ลิตร)\n✅ มีตัวเลือกใดตัวหนึ่งถูกเลือกอัตโนมัติ"),

    ("TC005.4", "Unit Filtering", "กรองหน่วยสำหรับกาว (556)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กาว (556)'\n2. คลิก dropdown 'หน่วยหลัก'",
     "✅ แสดงเฉพาะหน่วยหมึก (กระป๋อง, กิโล)"),

    ("TC005.5", "Unit Filtering", "หน่วยสำหรับกลุ่มอื่นๆ",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'เคลือบ (557)'\n2. คลิก dropdown 'หน่วยหลัก'\n3. นับจำนวนตัวเลือก",
     "✅ แสดงหน่วยทั้งหมด (ไม่กรอง)\n✅ จำนวนเท่ากับตอนไม่เลือกกลุ่ม"),

    ("TC005.6", "Unit Filtering", "เปลี่ยนกลุ่มหลังเลือกหน่วยแล้ว",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n2. เลือก 'หน่วยหลัก' = 'แผ่น'\n3. เปลี่ยน 'กลุ่มวัสดุ' เป็น 'หมึก (553)'\n4. ดู dropdown 'หน่วยหลัก'",
     "✅ รายการหน่วยเปลี่ยนเป็นหน่วยของหมึก\n✅ หน่วยที่เลือกไว้ถูก reset\n✅ ไม่มีค่า 'แผ่น' ให้เลือก"),

    # TC006: Auto-Naming
    ("TC006.1", "Auto-Naming", "ตั้งชื่ออัตโนมัติ - กรอกครบทุกฟิลด์",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n2. กรอก 'ประเภทกระดาษ (ไทย)' = 'กระดาษคาร์ตัน'\n3. กรอก 'แบรนด์ 1' = 'MOORIM'\n4. กรอก 'GSM' = '300'\n5. กรอก 'ความกว้าง (มม.)' = '840'\n6. กรอก 'ความยาว (มม.)' = '900'\n7. ดูฟิลด์ 'ชื่อวัสดุ'",
     "✅ ชื่อวัสดุถูกตั้งเป็น: 'กระดาษคาร์ตัน MOORIM 300g 33.07×35.43นิ้ว'\n✅ อัปเดตอัตโนมัติทันทีหลังกรอกฟิลด์สุดท้าย\n✅ ไม่สามารถแก้ไขได้ (readonly)"),

    ("TC006.2", "Auto-Naming", "ตั้งชื่ออัตโนมัติ - กรอกไม่ครบ",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n2. กรอก 'ประเภทกระดาษ (ไทย)' = 'กระดาษคาร์ตัน'\n3. กรอก 'แบรนด์ 1' = 'MOORIM'\n4. กรอก 'GSM' = '300'\n5. ปล่อย 'ความกว้าง' และ 'ความยาว' ว่าง\n6. ดูฟิลด์ 'ชื่อวัสดุ'",
     "✅ ชื่อวัสดุยังว่างอยู่\n✅ ไม่สร้างชื่อบางส่วน"),

    ("TC006.3", "Auto-Naming", "อัปเดตชื่อเมื่อแก้ไขข้อมูล",
     "ทำ TC006.1 แล้ว, มีชื่ออัตโนมัติแล้ว",
     "1. เปลี่ยน 'แบรนด์ 1' จาก 'MOORIM' เป็น 'STAR'\n2. ดูฟิลด์ 'ชื่อวัสดุ'",
     "✅ ชื่อเปลี่ยนเป็น: 'กระดาษคาร์ตัน STAR 300g 33.07×35.43นิ้ว'\n✅ อัปเดตทันที"),

    ("TC006.4", "Auto-Naming", "อัปเดตชื่อเมื่อเปลี่ยน GSM",
     "ทำ TC006.1 แล้ว",
     "1. เปลี่ยน 'GSM' จาก '300' เป็น '350'\n2. ดูฟิลด์ 'ชื่อวัสดุ'",
     "✅ ชื่อเปลี่ยนเป็น: 'กระดาษคาร์ตัน MOORIM 350g 33.07×35.43นิ้ว'"),

    ("TC006.5", "Auto-Naming", "อัปเดตชื่อเมื่อเปลี่ยนขนาด",
     "ทำ TC006.1 แล้ว",
     "1. เปลี่ยน 'ความกว้าง (มม.)' จาก '840' เป็น '1000'\n2. ดูฟิลด์ 'ชื่อวัสดุ'",
     "✅ ขนาดนิ้วคำนวณใหม่: 1000 ÷ 25.4 = 39.37 นิ้ว\n✅ ชื่อเปลี่ยนเป็น: 'กระดาษคาร์ตัน MOORIM 300g 39.37×35.43นิ้ว'"),

    ("TC006.6", "Auto-Naming", "ชื่อกลับไปเป็นกรอกได้เมื่อเปลี่ยนกลุ่ม",
     "ทำ TC006.1 แล้ว, มีชื่ออัตโนมัติ",
     "1. เปลี่ยน 'กลุ่มวัสดุ' เป็น 'หมึก (553)'\n2. คลิกที่ฟิลด์ 'ชื่อวัสดุ'",
     "✅ ฟิลด์ 'ชื่อวัสดุ' กรอกได้ (ไม่เป็น readonly)\n✅ พื้นหลังไม่เป็นสีเทา\n✅ สามารถแก้ไขชื่อได้"),

    # TC007: Conversion
    ("TC007.1", "Conversion", "แปลง mm → inch (กว้าง)",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก 'ความกว้าง (มม.)' = '840'\n2. Tab หรือคลิกออกจากฟิลด์\n3. ดูฟิลด์ 'ความกว้าง (นิ้ว)'",
     "✅ แสดง: '33.07' (840 ÷ 25.4 = 33.07)\n✅ ทศนิยม 2 ตำแหน่ง\n✅ อัปเดตทันที"),

    ("TC007.2", "Conversion", "แปลง mm → inch (ยาว)",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก 'ความยาว (มม.)' = '900'\n2. ดูฟิลด์ 'ความยาว (นิ้ว)'",
     "✅ แสดง: '35.43' (900 ÷ 25.4 = 35.43)"),

    ("TC007.3", "Conversion", "แปลง inch → mm (กว้าง)",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก 'ความกว้าง (นิ้ว)' = '33'\n2. Tab หรือคลิกออกจากฟิลด์\n3. ดูฟิลด์ 'ความกว้าง (มม.)'",
     "✅ แสดง: '838.20' (33 × 25.4 = 838.20)\n✅ ทศนิยม 2 ตำแหน่ง"),

    ("TC007.4", "Conversion", "แปลง inch → mm (ยาว)",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก 'ความยาว (นิ้ว)' = '35'\n2. ดูฟิลด์ 'ความยาว (มม.)'",
     "✅ แสดง: '889.00' (35 × 25.4 = 889.00)"),

    ("TC007.5", "Conversion", "ทศนิยมความละเอียดสูง",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก 'ความกว้าง (มม.)' = '25.4'\n2. ดูฟิลด์ 'ความกว้าง (นิ้ว)'",
     "✅ แสดง: '1.00' (25.4 ÷ 25.4 = 1.00)"),

    ("TC007.6", "Conversion", "กรอกค่า 0 หรือว่าง",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก 'ความกว้าง (มม.)' = '0'\n2. ดูฟิลด์ 'ความกว้าง (นิ้ว)'\n3. ลบค่าให้เป็นว่าง\n4. ดูฟิลด์ 'ความกว้าง (นิ้ว)'",
     "✅ ขั้นที่ 2: แสดงว่าง (ไม่แสดง '0.00')\n✅ ขั้นที่ 4: แสดงว่าง"),

    # TC008: Weight Calculation
    ("TC008.1", "Weight Calculation", "คำนวณน้ำหนัก - ตัวเลขปกติ",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก 'ความกว้าง (มม.)' = '840'\n2. กรอก 'ความยาว (มม.)' = '900'\n3. กรอก 'GSM' = '300'\n4. ดูฟิลด์ 'น้ำหนักต่อแผ่น (กก.)'\n5. ดูข้อความ Calculation Display",
     "✅ น้ำหนัก = 0.2268 กก. (สูตร: (840 × 900 × 300) ÷ 1,000,000 = 0.2268)\n✅ ทศนิยม 4 ตำแหน่ง\n✅ Calculation Display แสดง: '840 × 900 × 300 ÷ 1,000,000 = 0.2268 กก./แผ่น'\n✅ ตัวเลขเป็นสีเขียว (success)"),

    ("TC008.2", "Weight Calculation", "คำนวณน้ำหนัก - กรณีทดสอบที่ 2",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก W=1000, L=1000, GSM=250\n2. ดูผลลัพธ์",
     "✅ น้ำหนัก = 0.2500 กก. (สูตร: (1000 × 1000 × 250) ÷ 1,000,000 = 0.2500)"),

    ("TC008.3", "Weight Calculation", "คำนวณน้ำหนัก - กรณีทดสอบที่ 3",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอก W=500, L=600, GSM=200\n2. ดูผลลัพธ์",
     "✅ น้ำหนัก = 0.0600 กก. (สูตร: (500 × 600 × 200) ÷ 1,000,000 = 0.0600)"),

    ("TC008.4", "Weight Calculation", "ไม่คำนวณเมื่อกรอกไม่ครบ",
     "เลือกกลุ่มกระดาษแล้ว",
     "1. กรอกเฉพาะ 'ความกว้าง (มม.)' = '840'\n2. ปล่อย 'ความยาว' และ 'GSM' ว่าง\n3. ดูฟิลด์ 'น้ำหนักต่อแผ่น'",
     "✅ น้ำหนักว่าง (ไม่แสดงค่า)\n✅ Calculation Display ว่าง หรือแสดงข้อความ hint"),

    ("TC008.5", "Weight Calculation", "อัปเดตเมื่อเปลี่ยนค่า",
     "ทำ TC008.1 แล้ว, มีน้ำหนักแสดงอยู่",
     "1. จดน้ำหนักเดิม (0.2268)\n2. เปลี่ยน 'GSM' จาก '300' เป็น '350'\n3. ดูน้ำหนักใหม่",
     "✅ น้ำหนักเปลี่ยนเป็น 0.2646 กก. ((840 × 900 × 350) ÷ 1,000,000 = 0.2646)\n✅ อัปเดตทันทีโดยไม่ต้องคลิกปุ่มใดๆ"),

    # TC009: Foil Area Calculation
    ("TC009.1", "Foil Area Calc", "คำนวณพื้นที่ฟอยล์ - กรณีปกติ",
     "เลือกกลุ่มฟอยล์ (555) แล้ว",
     "1. กรอก 'ความกว้าง (มม.)' = '500'\n2. กรอก 'ความยาว (เมตร)' = '100'\n3. ดูฟิลด์ 'พื้นที่ (ตร.ม.)'",
     "✅ พื้นที่ = 50.0000 ตร.ม. (สูตร: (500 ÷ 1000) × 100 = 0.5 × 100 = 50.0000)\n✅ ทศนิยม 4 ตำแหน่ง\n✅ Readonly (ไม่สามารถแก้ไขได้)"),

    ("TC009.2", "Foil Area Calc", "คำนวณพื้นที่ฟอยล์ - กรณีทดสอบที่ 2",
     "เลือกกลุ่มฟอยล์แล้ว",
     "1. กรอก W=1000 มม., L=50 ม.\n2. ดูผลลัพธ์",
     "✅ พื้นที่ = 50.0000 ตร.ม. ((1000 ÷ 1000) × 50 = 1 × 50 = 50.0000)"),

    ("TC009.3", "Foil Area Calc", "ไม่คำนวณเมื่อกรอกไม่ครบ",
     "เลือกกลุ่มฟอยล์แล้ว",
     "1. กรอกเฉพาะ 'ความกว้าง' = '500'\n2. ปล่อย 'ความยาว' ว่าง\n3. ดูฟิลด์ 'พื้นที่'",
     "✅ พื้นที่ว่าง (ไม่แสดงค่า)"),

    # TC010: Select2
    ("TC010.1", "Select2 Search", "ค้นหาด้วยชื่อผู้ขาย",
     "โหลดหน้าเพจแล้ว, มีผู้ขายชื่อ 'ABC Company (001)'",
     "1. คลิก dropdown 'ผู้ขาย'\n2. พิมพ์ 'ABC' ในช่องค้นหา\n3. ดูผลลัพธ์ที่แสดง",
     "✅ แสดงเฉพาะผู้ขายที่มีคำว่า 'ABC' ในชื่อ\n✅ ไม่แสดงผู้ขายอื่นๆ\n✅ ค้นหาแบบ case-insensitive"),

    ("TC010.2", "Select2 Search", "ค้นหาด้วยรหัสผู้ขาย",
     "โหลดหน้าเพจแล้ว, มีผู้ขาย 'ABC Company (001)'",
     "1. คลิก dropdown 'ผู้ขาย'\n2. พิมพ์ '001' ในช่องค้นหา\n3. ดูผลลัพธ์ที่แสดง",
     "✅ แสดงผู้ขายที่มีรหัส '001'\n✅ ค้นหาจาก data-code ได้"),

    ("TC010.3", "Select2 Search", "ค้นหาไม่เจอ",
     "โหลดหน้าเพจแล้ว",
     "1. คลิก dropdown 'ผู้ขาย'\n2. พิมพ์ 'ZZZZZZZZZ'\n3. ดูผลลัพธ์",
     "✅ แสดงข้อความ 'No results found' หรือเทียบเท่า\n✅ ไม่แสดงรายการใดๆ"),

    ("TC010.4", "Select2 Search", "Clear Selection",
     "เลือกผู้ขายแล้ว",
     "1. เลือก 'ผู้ขาย' = 'ABC Company'\n2. คลิกปุ่ม X (clear) ที่ Select2\n3. ดู dropdown",
     "✅ ค่าถูกล้าง\n✅ แสดง placeholder 'เลือกหรือค้นหาผู้ขาย...'\n✅ SSP Preview กลับไปเป็น '-----'"),

    ("TC010.5", "Select2 Search", "เปิด Dropdown",
     "โหลดหน้าเพจแล้ว",
     "1. คลิก dropdown 'ผู้ขาย'\n2. ดูรายการที่แสดง",
     "✅ แสดงผู้ขายทั้งหมด (ไม่กรอง)\n✅ มีช่องค้นหาด้านบน\n✅ แสดงลูกศรลง/ขึ้นตามสถานะ"),

    # TC011: Validation
    ("TC011.1", "Form Validation", "Submit โดยไม่กรอกอะไรเลย",
     "โหลดหน้าเพจแล้ว",
     "1. คลิกปุ่ม 'บันทึกข้อมูล' ทันที",
     "✅ แสดง Alert 'กรุณากรอกข้อมูลที่จำเป็นให้ครบถ้วน'\n✅ ไม่ส่งข้อมูลไป server\n✅ ไม่เรียก API generate_ssp_code"),

    ("TC011.2", "Form Validation", "Submit กรอกไม่ครบ (ไม่มีประเภท)",
     "โหลดหน้าเพจแล้ว",
     "1. ปล่อย 'ประเภทวัสดุ' ว่าง\n2. เลือก 'กลุ่มวัสดุ' = 'กระดาษคาร์ตัน (501)'\n3. เลือก 'ผู้ขาย' = 'ABC Company'\n4. กรอก 'ชื่อวัสดุ'\n5. คลิก 'บันทึกข้อมูล'",
     "✅ แสดง Alert 'กรุณากรอกข้อมูลที่จำเป็นให้ครบถ้วน'\n✅ ไม่ส่งข้อมูล"),

    ("TC011.3", "Form Validation", "Submit กรอกไม่ครบ (ไม่มีกลุ่ม)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'ประเภทวัสดุ' = 'Raw Material'\n2. ปล่อย 'กลุ่มวัสดุ' ว่าง\n3. เลือก 'ผู้ขาย'\n4. กรอก 'ชื่อวัสดุ'\n5. คลิก 'บันทึกข้อมูล'",
     "✅ แสดง Alert 'กรุณากรอกข้อมูลที่จำเป็นให้ครบถ้วน'"),

    ("TC011.4", "Form Validation", "Submit กรอกไม่ครบ (ไม่มีผู้ขาย)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'ประเภทวัสดุ'\n2. เลือก 'กลุ่มวัสดุ'\n3. ปล่อย 'ผู้ขาย' ว่าง\n4. กรอก 'ชื่อวัสดุ'\n5. คลิก 'บันทึกข้อมูล'",
     "✅ แสดง Alert 'กรุณากรอกข้อมูลที่จำเป็นให้ครบถ้วน'"),

    ("TC011.5", "Form Validation", "Submit กรอกไม่ครบ (ไม่มีชื่อวัสดุ)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'ประเภทวัสดุ'\n2. เลือก 'กลุ่มวัสดุ' = 'หมึก (553)'\n3. เลือก 'ผู้ขาย'\n4. ปล่อย 'ชื่อวัสดุ' ว่าง\n5. คลิก 'บันทึกข้อมูล'",
     "✅ แสดง Alert 'กรุณากรอกข้อมูลที่จำเป็นให้ครบถ้วน'"),

    ("TC011.6", "Form Validation", "Submit สำเร็จ (กรอกครบ)",
     "โหลดหน้าเพจแล้ว",
     "1. เลือก 'ประเภทวัสดุ' = 'Raw Material (R)'\n2. เลือก 'กลุ่มวัสดุ' = 'หมึก (553)'\n3. เลือก 'ผู้ขาย' = 'ABC Company (001)'\n4. เลือก 'หน่วยหลัก' = 'กระป๋อง'\n5. กรอก 'ชื่อวัสดุ' = 'Test Ink'\n6. คลิก 'บันทึกข้อมูล'\n7. ดู Network tab",
     "✅ ไม่แสดง Alert error\n✅ เรียก API POST generate_ssp_code.php\n✅ ได้ SSP Code กลับมา\n✅ เพิ่ม hidden input ssp_code ใน form\n✅ Submit form ไป save_material.php\n✅ Redirect หรือแสดงผลสำเร็จ"),

    # TC012: Form Submission
    ("TC012.1", "Form Submission", "Generate SSP Code ก่อน Submit",
     "กรอกข้อมูลครบถ้วนแล้ว",
     "1. กรอกฟอร์มครบ\n2. เปิด Network tab\n3. คลิก 'บันทึกข้อมูล'\n4. ดูคำขอที่ส่งไป",
     "✅ ส่งคำขอ POST ไป generate_ssp_code.php ก่อน\n✅ Payload มี: material_type_id, group_id, supplier_id\n✅ ได้ Response กลับมา: {success: true, ssp_code: 'R501001XXXXX'}\n✅ จากนั้นถึง submit ไป save_material.php"),

    ("TC012.2", "Form Submission", "Hidden Input SSP Code",
     "กรอกข้อมูลครบ, Generate SSP Code สำเร็จ",
     "1. กรอกฟอร์มครบ\n2. คลิก 'บันทึกข้อมูล'\n3. รอ API generate SSP Code\n4. เปิด Console พิมพ์: $('input[name=\"ssp_code\"]').val()",
     "✅ Console แสดง SSP Code ที่ได้จาก API\n✅ มี hidden input ใน form\n✅ ค่าตรงกับที่ API ส่งกลับมา"),

    ("TC012.3", "Form Submission", "API Generate SSP ล้มเหลว",
     "ปิด API server หรือแก้ URL API ให้ผิด",
     "1. กรอกฟอร์มครบ\n2. คลิก 'บันทึกข้อมูล'\n3. ดู Alert/Console",
     "✅ แสดง Alert 'เกิดข้อผิดพลาดในการสร้าง SSP Code'\n✅ ไม่ส่งข้อมูลไป save_material.php\n✅ Console แสดง error"),

    ("TC012.4", "Form Submission", "ข้อมูลที่ส่งไป save_material.php",
     "กรอกฟอร์มครบ (กระดาษ)",
     "1. เลือกกลุ่มกระดาษ\n2. กรอกข้อมูลครบ\n3. คลิก 'บันทึกข้อมูล'\n4. ดู Network > save_material.php > Payload",
     "✅ มีฟิลด์: ssp_code (hidden), material_type_id, group_id, supplier_id, unit_id, name, name2, type_paperboard_th, brand, brand2, gsm, caliper, w_mm, l_mm, w_inch, l_inch, weight_kg_per_sheet, laminated1, laminated2, certificated"),

    # TC013: Preview Modal
    ("TC013.1", "Preview Modal", "เปิด Preview Modal",
     "กรอกข้อมูลแล้ว (กระดาษ)",
     "1. กรอกข้อมูลครบ\n2. คลิกปุ่ม 'ดูตัวอย่าง'\n3. ดู modal ที่เปิดขึ้น",
     "✅ เปิด modal 'ตัวอย่างข้อมูลที่จะบันทึก'\n✅ แสดงหัวข้อ 'ข้อมูลหลัก'\n✅ แสดง SSP Code, ชื่อวัสดุ, ประเภท, กลุ่ม, ผู้ขาย, หน่วย\n✅ แสดงหัวข้อ 'รายละเอียดกระดาษ'\n✅ แสดงแบรนด์, GSM, ขนาด, น้ำหนัก\n✅ มีปุ่ม 'ปิด' และ 'ยืนยันและบันทึก'"),

    ("TC013.2", "Preview Modal", "ยืนยันจาก Modal",
     "เปิด Preview Modal แล้ว",
     "1. เปิด modal preview\n2. คลิกปุ่ม 'ยืนยันและบันทึก'\n3. ดู Network tab",
     "✅ ปิด modal\n✅ Submit form (เหมือน TC012)\n✅ เรียก API generate SSP Code\n✅ ส่งข้อมูลไป save_material.php"),

    ("TC013.3", "Preview Modal", "ปิด Modal",
     "เปิด Preview Modal แล้ว",
     "1. เปิด modal preview\n2. คลิกปุ่ม 'ปิด' หรือ X\n3. ดูหน้าเพจ",
     "✅ ปิด modal\n✅ กลับมาที่ฟอร์ม\n✅ ข้อมูลยังอยู่ครบ (ไม่หาย)\n✅ ไม่ submit"),

    # TC014: Edge Cases
    ("TC014.1", "Edge Cases", "ชื่อวัสดุมีอักขระพิเศษ",
     "เลือกกลุ่มหมึก (ชื่อกรอกเองได้)",
     "1. กรอก 'ชื่อวัสดุ' = 'Test ~!@#$%^&*()_+-={}[]'\n2. Submit form",
     "✅ บันทึกสำเร็จ\n✅ ไม่มี error\n✅ ข้อมูลถูกต้องใน database"),

    ("TC014.2", "Edge Cases", "ตัวเลขทศนิยมหลายหลัก",
     "เลือกกลุ่มกระดาษ",
     "1. กรอก 'GSM' = '123.456789'\n2. กรอก 'Caliper' = '0.123456789'\n3. ดูการคำนวณน้ำหนัก",
     "✅ รับค่าได้\n✅ คำนวณถูกต้อง\n✅ แสดงทศนิยมตามที่กำหนด (4 ตำแหน่ง)"),

    ("TC014.3", "Edge Cases", "ตัวเลขติดลบ",
     "เลือกกลุ่มกระดาษ",
     "1. กรอก 'ความกว้าง (มม.)' = '-100'\n2. ดูฟิลด์ 'ความกว้าง (นิ้ว)' และน้ำหนัก",
     "✅ ไม่ยอมให้กรอก (HTML5 validation) หรือ ไม่คำนวณ (แสดงว่าง)"),

    ("TC014.4", "Edge Cases", "ตัวเลขใหญ่มาก",
     "เลือกกลุ่มกระดาษ",
     "1. กรอก 'ความกว้าง' = '999999999'\n2. กรอก 'ความยาว' = '999999999'\n3. กรอก 'GSM' = '999999'\n4. ดูน้ำหนัก",
     "✅ คำนวณได้ (แม้ค่าจะไม่สมเหตุสมผล)\n✅ ไม่มี error\n✅ แสดงผลถูกต้อง"),

    ("TC014.5", "Edge Cases", "ช่องว่างนำหน้า/ท้าย",
     "เลือกกลุ่มหมึก",
     "1. กรอก 'ชื่อวัสดุ' = '  Test Material  ' (มีช่องว่างหน้าท้าย)\n2. Submit\n3. ดูข้อมูลที่บันทึก",
     "✅ ระบบควร trim ช่องว่าง (backend)\n✅ บันทึกเป็น 'Test Material' (ไม่มีช่องว่าง)"),

    ("TC014.6", "Edge Cases", "เปลี่ยนกลุ่มหลังกรอกข้อมูล",
     "โหลดหน้าเพจแล้ว",
     "1. เลือกกลุ่มกระดาษ\n2. กรอกข้อมูลกระดาษครบ\n3. เปลี่ยนเป็นกลุ่มหมึก\n4. ดูข้อมูลที่กรอก",
     "✅ ข้อมูลหลักยังอยู่ (Type, Group, Supplier, Unit)\n✅ ชื่อวัสดุที่สร้างอัตโนมัติยังอยู่\n✅ ข้อมูลเฉพาะกระดาษหาย (ไม่ส่งไปด้วย)\n✅ แสดงฟอร์มหมึกแทน"),

    ("TC014.7", "Edge Cases", "Refresh หน้าระหว่างกรอก",
     "กรอกข้อมูลครึ่งหนึ่งแล้ว",
     "1. กรอกข้อมูลบางส่วน\n2. กด F5 หรือ Refresh\n3. ดูฟอร์ม",
     "✅ ข้อมูลหายหมด (reset)\n✅ กลับไปสถานะเริ่มต้น\n✅ ไม่มี error"),

    # TC015: Browser Compatibility
    ("TC015.1", "Browser Compat", "Chrome",
     "ติดตั้ง Chrome เวอร์ชันล่าสุด",
     "1. เปิดหน้าเพจใน Chrome\n2. ทดสอบ TC001-TC012",
     "✅ ทุก feature ทำงานปกติ\n✅ ไม่มี error ใน Console"),

    ("TC015.2", "Browser Compat", "Firefox",
     "ติดตั้ง Firefox เวอร์ชันล่าสุด",
     "1. เปิดหน้าเพจใน Firefox\n2. ทดสอบ TC001-TC012",
     "✅ ทุก feature ทำงานปกติ"),

    ("TC015.3", "Browser Compat", "Safari",
     "ใช้ macOS + Safari",
     "1. เปิดหน้าเพจใน Safari\n2. ทดสอบ TC001-TC012",
     "✅ ทุก feature ทำงานปกติ"),

    ("TC015.4", "Browser Compat", "Edge",
     "ติดตั้ง Edge เวอร์ชันล่าสุด",
     "1. เปิดหน้าเพจใน Edge\n2. ทดสอบ TC001-TC012",
     "✅ ทุก feature ทำงานปกติ"),

    # TC016: Responsive Design
    ("TC016.1", "Responsive", "Desktop (1920×1080)",
     "ใช้ desktop หรือ resize browser",
     "1. ตั้งขนาดหน้าจอ 1920×1080\n2. ดูเลย์เอาท์\n3. ทดสอบฟังก์ชัน",
     "✅ Layout สวยงาม ไม่มีส่วนเกิน\n✅ ฟอร์มแสดงเต็ม\n✅ ทุกฟังก์ชันทำงาน"),

    ("TC016.2", "Responsive", "Laptop (1366×768)",
     "Resize browser เป็น 1366×768",
     "1. ตั้งขนาดหน้าจอ 1366×768\n2. ดูเลย์เอาท์",
     "✅ Layout ปรับตัว\n✅ ไม่มี horizontal scroll"),

    ("TC016.3", "Responsive", "Tablet (768×1024)",
     "Resize browser หรือใช้ device emulator",
     "1. ตั้งขนาดหน้าจอ 768×1024\n2. ทดสอบฟอร์ม",
     "✅ Layout stack เป็นคอลัมน์เดียว\n✅ Select2 ทำงานได้\n✅ ปุ่มกดได้ง่าย"),

    ("TC016.4", "Responsive", "Mobile (375×667)",
     "Resize browser หรือใช้มือถือจริง",
     "1. ตั้งขนาดหน้าจอ 375×667 (iPhone SE)\n2. ทดสอบฟอร์ม",
     "✅ Layout responsive\n✅ ฟอร์มกรอกได้\n✅ Select2 เปิดได้\n✅ ปุ่มไม่ซ้อนกัน"),
]

# Add test cases to worksheet
current_row = 2
for tc in test_cases:
    ws.append([tc[0], tc[1], tc[2], tc[3], tc[4], tc[5], "", "", ""])

    # Apply borders and alignment
    for col_num in range(1, 10):
        cell = ws.cell(row=current_row, column=col_num)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    current_row += 1

# Set column widths
ws.column_dimensions['A'].width = 12  # TC ID
ws.column_dimensions['B'].width = 18  # Category
ws.column_dimensions['C'].width = 35  # Description
ws.column_dimensions['D'].width = 30  # Pre-conditions
ws.column_dimensions['E'].width = 50  # Test Steps
ws.column_dimensions['F'].width = 60  # Expected Result
ws.column_dimensions['G'].width = 30  # Actual Result
ws.column_dimensions['H'].width = 12  # Status
ws.column_dimensions['I'].width = 25  # Notes

# Set row heights
ws.row_dimensions[1].height = 30
for row in range(2, current_row):
    ws.row_dimensions[row].height = 80

# Add validation for Status column
from openpyxl.worksheet.datavalidation import DataValidation

status_validation = DataValidation(
    type="list",
    formula1='"Pass,Fail,Pending,N/A"',
    allow_blank=True
)
status_validation.error = 'Invalid status'
status_validation.errorTitle = 'Invalid Entry'
status_validation.prompt = 'Please select status'
status_validation.promptTitle = 'Status'

ws.add_data_validation(status_validation)
status_validation.add(f'H2:H{current_row-1}')

# Create Summary sheet
summary_ws = wb.create_sheet("Summary", 0)

summary_data = [
    ["Manual Test Cases - pages/materials/add.php"],
    [""],
    ["Project:", "Production System"],
    ["Module:", "Materials Management"],
    ["Page:", "pages/materials/add.php"],
    ["Test Date:", ""],
    ["Tester:", ""],
    ["Browser:", ""],
    ["Environment:", ""],
    [""],
    ["Test Summary"],
    ["Category", "Total", "Passed", "Failed", "Pending", "Pass Rate"],
    ["Initial Page Load", 5, "", "", "", ""],
    ["SSP Code Preview", 5, "", "", "", ""],
    ["Form Visibility", 10, "", "", "", ""],
    ["Unit Filtering", 6, "", "", "", ""],
    ["Auto-Naming", 6, "", "", "", ""],
    ["Conversion", 6, "", "", "", ""],
    ["Weight Calculation", 5, "", "", "", ""],
    ["Foil Area Calc", 3, "", "", "", ""],
    ["Select2 Search", 5, "", "", "", ""],
    ["Form Validation", 6, "", "", "", ""],
    ["Form Submission", 4, "", "", "", ""],
    ["Preview Modal", 3, "", "", "", ""],
    ["Edge Cases", 7, "", "", "", ""],
    ["Browser Compat", 4, "", "", "", ""],
    ["Responsive", 4, "", "", "", ""],
    ["TOTAL", 79, "", "", "", ""],
]

for row_data in summary_data:
    summary_ws.append(row_data)

# Style summary sheet
summary_ws.merge_cells('A1:F1')
title_cell = summary_ws['A1']
title_cell.font = Font(bold=True, size=16, color="FFFFFF")
title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
summary_ws.row_dimensions[1].height = 30

# Style info section
for row in range(3, 10):
    summary_ws.cell(row=row, column=1).font = Font(bold=True)

# Style summary table header
summary_ws.merge_cells('A11:F11')
summary_header = summary_ws['A11']
summary_header.font = Font(bold=True, size=14, color="FFFFFF")
summary_header.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
summary_header.alignment = Alignment(horizontal="center", vertical="center")

for col in range(1, 7):
    cell = summary_ws.cell(row=12, column=col)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Style summary data rows
for row in range(13, 29):
    for col in range(1, 7):
        cell = summary_ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Bold the category name
    summary_ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # Add formula for pass rate (column F)
    if row < 28:  # Not the TOTAL row
        summary_ws.cell(row=row, column=6).value = f'=IF(B{row}>0,C{row}/B{row}*100,0)&"%"'
    else:  # TOTAL row
        summary_ws.cell(row=row, column=1).font = Font(bold=True)
        summary_ws.cell(row=row, column=2).value = f'=SUM(B13:B27)'
        summary_ws.cell(row=row, column=3).value = f'=SUM(C13:C27)'
        summary_ws.cell(row=row, column=4).value = f'=SUM(D13:D27)'
        summary_ws.cell(row=row, column=5).value = f'=SUM(E13:E27)'
        summary_ws.cell(row=row, column=6).value = f'=IF(B28>0,C28/B28*100,0)&"%"'

# Set column widths for summary
summary_ws.column_dimensions['A'].width = 25
summary_ws.column_dimensions['B'].width = 12
summary_ws.column_dimensions['C'].width = 12
summary_ws.column_dimensions['D'].width = 12
summary_ws.column_dimensions['E'].width = 12
summary_ws.column_dimensions['F'].width = 15

# Add Notes sheet
notes_ws = wb.create_sheet("Notes")
notes_ws['A1'] = "Testing Notes & Instructions"
notes_ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
notes_ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
notes_ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
notes_ws.merge_cells('A1:D1')
notes_ws.row_dimensions[1].height = 25

notes_content = [
    [""],
    ["Important Notes:"],
    ["1. ทดสอบใน environment ที่มีข้อมูล master data ครบถ้วน"],
    ["2. ตรวจสอบ Console errors ทุกครั้ง"],
    ["3. บันทึก screenshot สำหรับ failed cases"],
    ["4. Test ใน database แยก (ไม่ใช่ production)"],
    ["5. ตรวจสอบ Network tab สำหรับ API calls"],
    [""],
    ["Test Status Legend:"],
    ["Pass - Test case ผ่านตามที่คาดหวัง"],
    ["Fail - Test case ไม่ผ่าน พบข้อผิดพลาด"],
    ["Pending - ยังไม่ได้ทดสอบ"],
    ["N/A - ไม่สามารถทดสอบได้ (Not Applicable)"],
    [""],
    ["Expected Test Data:"],
    ["- ประเภทวัสดุ: Raw Material (R), Finished Goods (F)"],
    ["- กลุ่มวัสดุ: กระดาษคาร์ตัน (501), หมึก (553), กาว (556), ฟอยล์ (555), etc."],
    ["- ผู้ขาย: ABC Company (001), XYZ Supplier (002), etc."],
    ["- หน่วย: แผ่น, กระป๋อง, กิโลกรัม, etc."],
    [""],
    ["Test Formulas:"],
    ["- mm to inch: inch = mm ÷ 25.4"],
    ["- inch to mm: mm = inch × 25.4"],
    ["- น้ำหนักกระดาษ: (กว้าง × ยาว × GSM) ÷ 1,000,000 = กก./แผ่น"],
    ["- พื้นที่ฟอยล์: (กว้าง_มม. ÷ 1000) × ยาว_ม. = ตร.ม."],
]

for row_idx, row_data in enumerate(notes_content, start=2):
    notes_ws.append(row_data)
    if row_idx in [2, 9, 15, 20]:  # Header rows
        notes_ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)

notes_ws.column_dimensions['A'].width = 80

# Freeze panes in Test Cases sheet
ws.freeze_panes = 'A2'

# Save workbook
filename = '/home/user/Production/Manual_Test_Cases_Materials_Add.xlsx'
wb.save(filename)
print(f"✅ Excel file created successfully: {filename}")
print(f"📊 Total test cases: {len(test_cases)}")
print(f"📄 Sheets: Summary, Test Cases, Notes")
